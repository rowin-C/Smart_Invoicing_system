import argparse
import csv
import datetime
import os
import platform
import subprocess
import sys
from pathlib import Path
import threading
import uuid
from qrcodegen import *
from changeImage import duplicate_and_replace

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle




import torch

from datetime import datetime

now = datetime.now()

FILE = Path(__file__).resolve()
ROOT = FILE.parents[0]  # YOLOv5 root directory
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))  # add ROOT to PATH
ROOT = Path(os.path.relpath(ROOT, Path.cwd()))  # relative

from ultralytics.utils.plotting import Annotator, colors, save_one_box

from models.common import DetectMultiBackend
from utils.dataloaders import IMG_FORMATS, VID_FORMATS, LoadImages, LoadScreenshots, LoadStreams
from utils.general import (
    LOGGER,
    Profile,
    check_file,
    check_img_size,
    check_imshow,
    check_requirements,
    colorstr,
    cv2,
    increment_path,
    non_max_suppression,
    print_args,
    scale_boxes,
    strip_optimizer,
    xyxy2xywh,
)
from utils.torch_utils import select_device, smart_inference_mode
from flask import Flask, request, jsonify
from flask import Flask,  request


import sys
import pandas as pd
import cv2 
from pyzbar.pyzbar import decode
from flask_cors import CORS

import os
import signal
import sys


app = Flask(__name__)

# Enable CORS
CORS(app, resources={r"/*": {"origins": "http://localhost:3000"}})



def shutdown_server():
    print("Shutting down server gracefully...")
    os.kill(os.getpid(), signal.SIGINT)
    
    

    
#invoice gen
def convert_excel_to_pdf(excel_file, pdf_file):
    # Load Excel workbook
    wb = load_workbook(excel_file)
    sheet = wb.active

    # Extract data from Excel sheet
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    # Create PDF
    doc = SimpleDocTemplate(pdf_file, pagesize=letter)
    table = Table(data)

    # Build PDF
    doc.build([table])






@app.route("/opencam", methods=['POST'])
def opencam():
   
    data = request.json
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    value = run(**data)
         
    return value

@app.route("/detect", methods=['GET'])
def detect():
    BarcodedItem = False
    
    if(prev_detections == "barcodes"):
        frames = cv2.imread("code.png")
        cv2.imshow("frame", frames)
        detectedBarcode = decode(frames)
        # if no any barcode detected 
        if  detectedBarcode:
            for barcode in detectedBarcode:
                # if barcode is not blank 
                if barcode.data != "":
                    # Search for the barcode in the Excel file
                    # price = search_excel_file(barcode.data.decode('utf-8'))
                    # get name and price from the excel file
                    name, price = search_excel_file(barcode.data.decode('utf-8')) 
                    BarcodedItem = True
                    
                    return jsonify({'detection': name, 'price': str(price), 'BarcodedItem': BarcodedItem}), 200
    else:
        def get_price_from_excel(file_path, name):
            df = pd.read_excel(file_path)
            price = df.loc[df['name'] == name, 'price'].values[0]
            return price

        price = get_price_from_excel("Book1.xlsx", prev_detections)
        
        
    
    return jsonify({'detection': prev_detections, 'price': str(price), 'BarcodedItem': BarcodedItem}), 200



@app.route('/save', methods=['POST'])
def save():
    data = request.json
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    print(data)
# {'items': [{'detection': 'some', 'price': 12, 'BarcodedItem': True, 'value': 12}], 'totalcartvalue': 144, 'BillNo': 1311738448283418400, 'date': '6/3/2024'}    
   
    # Append data to Excel sheet
    
    BillNumber = data['BillNo']
    Time = data['date']
    
    items = data['items']
    totalcartvalue = data['totalcartvalue']
    
    # for item in items:
    #     item['BillNumber'] = BillNumber
    #     item['Time'] = now.strftime("%Y-%m-%d %H:%M:%S")
    
    for item in items:
        item['BillNumber'] = BillNumber
        item['Time'] = Time
        
       
    # # Append the data to the Excel file
    append_to_excel('Book2.xlsx', items)
    
    
    # convert_excel_to_pdf(f"Book2.xlsx", f"invoice_{BillNumber}.pdf")
    
    return 'Data saved', 200


    
@app.route('/stop', methods=['POST'])
def stop():
    IsRunning = False
    duplicate_and_replace("original.jpg", "code.jpg")
    shutdown_server()
    return 'Server shutting down...', 200



@app.route('/QRgen', methods=['POST'])
def QRgen():
    data = request.json
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    total_price = data['total_price']
    
    def to_svg_str(qr: QrCode, border: int) -> str:
        if border < 0:
            raise ValueError("Border must be non-negative")
        parts: List[str] = []
        for y in range(qr.get_size()):
            for x in range(qr.get_size()):
                if qr.get_module(x, y):
                    parts.append(f"M{x+border},{y+border}h1v1h-1z")
                    
        svgStr = f"""
<svg xmlns="http://www.w3.org/2000/svg" version="1.1" viewBox="0 0 {qr.get_size()+border*2} {qr.get_size()+border*2}" stroke="none">
	<rect width="100%" height="100%" fill="#FFFFFF"/>
	<path d="{" ".join(parts)}" fill="#000000"/>
</svg>
""" 
        
        return svgStr

            # Simple operation
    qr0 = QrCode.encode_text(f"upi://pay?pa=saubhagyaprasad31122001@oksbi&pn=Kami_%20Ronwinner&am={total_price}&cu=INR&aid=uGICAgICDuvWfXw", QrCode.Ecc.MEDIUM)
    svg = to_svg_str(qr0, 4)  # See qrcodegen-demo
    
    return svg
	    
            


# Append data to Excel sheet (remove the existing data and append the new data)

# def append_to_excel(existing_file, new_data):
#     try:
#         # Create a DataFrame from the new data
#         df_new = pd.DataFrame(new_data)

#         # Calculate the total price by multiplying each price by its quantity
#         total_price = (df_new['price'] * df_new['value']).sum()

#         # Write the new DataFrame to the Excel file, overwriting any existing content
#         with pd.ExcelWriter(existing_file, engine='openpyxl', mode='w') as writer:
#             df_new.to_excel(writer, index=False, sheet_name='Sheet1')

#             # Create a DataFrame for the total with a title and append it to the Excel file
#             df_total = pd.DataFrame({'Description': ['Total Price'], 'Amount': [total_price]})
#             df_total.to_excel(writer, startrow=df_new.shape[0]+3, index=False)

#             # Add a line on top of the total
#             worksheet = writer.sheets['Sheet1']
#             worksheet['A'+str(df_new.shape[0]+2)] = '-----------------------------'

#         print(f"Data written to '{existing_file}' successfully with total price.")

#     except FileNotFoundError:
#         print(f"Error: File '{existing_file}' not found.")
#     except Exception as e:
#         print(f"An error occurred: {e}")

def append_to_excel(existing_file, new_data):
    try:
        # Create a DataFrame from the new data
        df_new = pd.DataFrame(new_data)

        # Calculate the total price by multiplying each price by its quantity
        total_price = (df_new['price'] * df_new['value']).sum()

        # Read the existing data from the Excel file
        try:
            df_existing = pd.read_excel(existing_file, sheet_name='Sheet1')
        except FileNotFoundError:
            df_existing = pd.DataFrame()

        # Append the new data to the existing data
        df_combined = pd.concat([df_existing, df_new], ignore_index=True)

        # Write the combined DataFrame to the Excel file
        with pd.ExcelWriter(existing_file, engine='openpyxl', mode='w') as writer:
            df_combined.to_excel(writer, index=False, sheet_name='Sheet1')

            

        print(f"Data written to '{existing_file}' successfully with total price.")

    except Exception as e:
        print(f"An error occurred: {e}")





def search_excel_file(number):
    excel_file = 'Book1.xlsx'
    try:
        # Read the Excel file into a DataFrame
        df = pd.read_excel(excel_file)

        # Search for the number in the 'id' column
        result = df[df['id'] == int(number)]
        

        # Check if any matching row is found
        if not result.empty:
            # Print the details for the found number
            # return name and price
            return result["name"].values[0], result["price"].values[0]
        else:
            print(f"No details found for ID: {number}")
#if barcode is not found in the excel file then append the barcode in the excel file
            new_data = {
        'id': [int(number)],
        'name': "",
        'price': ""
         }
            
            append_to_excel(excel_file, new_data)
            
    except FileNotFoundError:
        print(f"Error: File '{excel_file}' not found in serach .")
    except Exception as e:
        print(f"An error occurred in search: {e}")




# YoloV5 detection code

prev_detections = "no detection"

@smart_inference_mode()
def run(
    weights=ROOT / "yolov5s.pt",  # model path or triton URL
    source=ROOT / "data/images",  # file/dir/URL/glob/screen/0(webcam)
    data=ROOT / "data/coco128.yaml",  # dataset.yaml path
    imgsz=(640, 640),  # inference size (height, width)
    conf_thres=0.25,  # confidence threshold
    iou_thres=0.45,  # NMS IOU threshold
    max_det=1000,  # maximum detections per image
    device="",  # cuda device, i.e. 0 or 0,1,2,3 or cpu
    view_img=False,  # show results
    save_txt=False,  # save results to *.txt
    save_csv=False,  # save results in CSV format
    save_conf=False,  # save confidences in --save-txt labels
    save_crop=False,  # save cropped prediction boxes
    nosave=False,  # do not save images/videos
    classes=None,  # filter by class: --class 0, or --class 0 2 3
    agnostic_nms=False,  # class-agnostic NMS
    augment=False,  # augmented inference
    visualize=False,  # visualize features
    update=False,  # update all models
    project=ROOT / "runs/detect",  # save results to project/name
    name="exp",  # save results to project/name
    exist_ok=False,  # existing project/name ok, do not increment
    line_thickness=3,  # bounding box thickness (pixels)
    hide_labels=False,  # hide labels
    hide_conf=False,  # hide confidences
    half=False,  # use FP16 half-precision inference
    dnn=False,  # use OpenCV DNN for ONNX inference
    vid_stride=1,  # video frame-rate stride
):
    global prev_detections 
    
    source = str(source)
    save_img = not nosave and not source.endswith(".txt")  # save inference images
    is_file = Path(source).suffix[1:] in (IMG_FORMATS + VID_FORMATS)
    is_url = source.lower().startswith(("rtsp://", "rtmp://", "http://", "https://"))
    webcam = source.isnumeric() or source.endswith(".streams") or (is_url and not is_file)
    screenshot = source.lower().startswith("screen")
    if is_url and is_file:
        source = check_file(source)  # download

    # Directories
    save_dir = increment_path(Path(project) / name, exist_ok=exist_ok)  # increment run
    (save_dir / "labels" if save_txt else save_dir).mkdir(parents=True, exist_ok=True)  # make dir
    (barDetect_dir := save_dir / "barDetect").mkdir(parents=True, exist_ok=True) 

    # Load model
    device = select_device(device)
    model = DetectMultiBackend(weights, device=device, dnn=dnn, data=data, fp16=half)
    stride, names, pt = model.stride, model.names, model.pt
    imgsz = check_img_size(imgsz, s=stride)  # check image size

    # Dataloader
    bs = 1  # batch_size
    if webcam:
        view_img = check_imshow(warn=True)
        dataset = LoadStreams(source, img_size=imgsz, stride=stride, auto=pt, vid_stride=vid_stride)
        bs = len(dataset)
    elif screenshot:
        dataset = LoadScreenshots(source, img_size=imgsz, stride=stride, auto=pt)
    else:
        dataset = LoadImages(source, img_size=imgsz, stride=stride, auto=pt, vid_stride=vid_stride)
    vid_path, vid_writer = [None] * bs, [None] * bs

    # Run inference
    model.warmup(imgsz=(1 if pt or model.triton else bs, 3, *imgsz))  # warmup
    seen, windows, dt = 0, [], (Profile(device=device), Profile(device=device), Profile(device=device))
    for path, im, im0s, vid_cap, s in dataset:
        with dt[0]:
            im = torch.from_numpy(im).to(model.device)
            im = im.half() if model.fp16 else im.float()  # uint8 to fp16/32
            im /= 255  # 0 - 255 to 0.0 - 1.0
            if len(im.shape) == 3:
                im = im[None]  # expand for batch dim
            if model.xml and im.shape[0] > 1:
                ims = torch.chunk(im, im.shape[0], 0)

        # Inference
        with dt[1]:
            visualize = increment_path(save_dir / Path(path).stem, mkdir=True) if visualize else False
            if model.xml and im.shape[0] > 1:
                pred = None
                for image in ims:
                    if pred is None:
                        pred = model(image, augment=augment, visualize=visualize).unsqueeze(0)
                    else:
                        pred = torch.cat((pred, model(image, augment=augment, visualize=visualize).unsqueeze(0)), dim=0)
                pred = [pred, None]
            else:
                pred = model(im, augment=augment, visualize=visualize)
        # NMS
        with dt[2]:
            pred = non_max_suppression(pred, conf_thres, iou_thres, classes, agnostic_nms, max_det=max_det)

        # Second-stage classifier (optional)
        # pred = utils.general.apply_classifier(pred, classifier_model, im, im0s)

        # Define the path for the CSV file
        csv_path = save_dir / "predictions.csv"

        # Create or append to the CSV file
        def write_to_csv(image_name, prediction, confidence):
            data = {"Image Name": image_name, "Prediction": prediction, "Confidence": confidence}
            with open(csv_path, mode="a", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=data.keys())
                if not csv_path.is_file():
                    writer.writeheader()
                writer.writerow(data)

        # Process predictions
        for i, det in enumerate(pred):  # per image
            seen += 1
            if webcam:  # batch_size >= 1
                p, im0, frame = path[i], im0s[i].copy(), dataset.count
                s += f"{i}: "
            else:
                p, im0, frame = path, im0s.copy(), getattr(dataset, "frame", 0)

            p = Path(p)  # to Path
            save_path = str(save_dir / p.name)  # im.jpg
            txt_path = str(save_dir / "labels" / p.stem) + ("" if dataset.mode == "image" else f"_{frame}")  # im.txt
            s += "%gx%g " % im.shape[2:]  # print string
            gn = torch.tensor(im0.shape)[[1, 0, 1, 0]]  # normalization gain whwh
            imc = im0.copy() if save_crop else im0  # for save_crop
            annotator = Annotator(im0, line_width=line_thickness, example=str(names))
            new_detections = set()

            if len(det):
                # Rescale boxes from img_size to im0 size
                det[:, :4] = scale_boxes(im.shape[2:], det[:, :4], im0.shape).round()

                # Print results
                for c in det[:, 5].unique():
                    n = (det[:, 5] == c).sum()  # detections per class
                    s += f"{n} {names[int(c)]}{'s' * (n > 1)}, "  # add to string

                # Write results
                for *xyxy, conf, cls in reversed(det):
                    c = int(cls)  # integer class
                    label = names[c] if hide_conf else f"{names[c]}"
                    confidence = float(conf)
                    confidence_str = f"{confidence:.2f}"

                    if save_csv:
                        write_to_csv(p.name, label, confidence_str)

                    if save_txt:  # Write to file
                        xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(-1).tolist()  # normalized xywh
                        line = (cls, *xywh, conf) if save_conf else (cls, *xywh)  # label format
                        with open(f"{txt_path}.txt", "a") as f:
                            f.write(("%g " * len(line)).rstrip() % line + "\n")

                    if save_img or save_crop or view_img:  # Add bbox to image
                        c = int(cls)  # integer class
                        label = None if hide_labels else (names[c] if hide_conf else f"{names[c]} {conf:.2f}")
                        annotator.box_label(xyxy, label, color=colors(c, True))
                    if save_crop:
                        save_one_box(xyxy, imc, file=save_dir / "crops" / names[c] / f"{p.stem}.jpg", BGR=True)
                    
                    # Save the image frame if the detected object is labeled as "barcodes"
                    if names[int(cls)]:
                     # Save the new image
                         cv2.imwrite(str("code.jpg"), im0)
                
                    
                    
                    new_detections.add((tuple(xyxy), c))

            # Check if new detections are different from previous ones
            if new_detections != prev_detections:
                # Stream results
                im0 = annotator.result()
                if view_img:
                    if platform.system() == "Linux" and p not in windows:
                        windows.append(p)
                        cv2.namedWindow(str(p), cv2.WINDOW_NORMAL | cv2.WINDOW_KEEPRATIO)  # allow window resize (Linux)
                        cv2.resizeWindow(str(p), im0.shape[1], im0.shape[0])
                    cv2.imshow(str(p), im0)
                    cv2.waitKey(1)  # 1 millisecond

                # Save results (image with detections)
                if save_img:
                    if dataset.mode == "image":
                        cv2.imwrite(save_path, im0)
                    else:  # 'video' or 'stream'
                        if vid_path[i] != save_path:  # new video
                            vid_path[i] = save_path
                            if isinstance(vid_writer[i], cv2.VideoWriter):
                                vid_writer[i].release()  # release previous video writer
                            if vid_cap:  # video
                                fps = vid_cap.get(cv2.CAP_PROP_FPS)
                                w = int(vid_cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                                h = int(vid_cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                            else:  # stream
                                fps, w, h = 30, im0.shape[1], im0.shape[0]
                            save_path = str(Path(save_path).with_suffix(".mp4"))  # force *.mp4 suffix on results videos
                            vid_writer[i] = cv2.VideoWriter(save_path, cv2.VideoWriter_fourcc(*"mp4v"), fps, (w, h))
                        vid_writer[i].write(im0)

                # Print time (inference-only)
                # LOGGER.info(f"{s}{'' if len(det) else '(no detections), '}{dt[1].dt * 1E3:.1f}ms")
                # print object name
                # if len(det):
                #     LOGGER.info(f"Object name: {names[int(det[0][5])]}")
                
                # if(len(det)):
                #     if(prev_detections == {names[int(det[0][5])]}):
                #         continue
                #     else:
                #         LOGGER.info(f"{names[int(det[0][5])]}")
                #         prev_detections = {names[int(det[0][5])]}
                
                    
                # yield  prev_detections
                if(len(det)):
                         if(prev_detections == names[int(det[0][5])]):
                             continue
                         else:
                            #  LOGGER.info(f"{names[int(det[0][5])]}")
                            #  yield names[int(det[0][5])]
                            #  print(prev_detections)
                            #  yield prev_detections
                             prev_detections = names[int(det[0][5])]

              

    # Print results
    t = tuple(x.t / seen * 1e3 for x in dt)  # speeds per image
    LOGGER.info(f"Speed: %.1fms pre-process, %.1fms inference, %.1fms NMS per image at shape {(1, 3, *imgsz)}" % t)
    if save_txt or save_img:
        s = f"\n{len(list(save_dir.glob('labels/*.txt')))} labels saved to {save_dir / 'labels'}" if save_txt else ""
        LOGGER.info(f"Results saved to {colorstr('bold', save_dir)}{s}")
    if update:
        strip_optimizer(weights[0])  # update model (to fix SourceChangeWarning)


def parse_opt():
    parser = argparse.ArgumentParser()
    parser.add_argument("--weights", nargs="+", type=str, default=ROOT / "yolov5s.pt", help="model path or triton URL")
    parser.add_argument("--source", type=str, default=ROOT / "data/images", help="file/dir/URL/glob/screen/0(webcam)")
    parser.add_argument("--data", type=str, default=ROOT / "data/coco128.yaml", help="(optional) dataset.yaml path")
    parser.add_argument("--imgsz", "--img", "--img-size", nargs="+", type=int, default=[640], help="inference size h,w")
    parser.add_argument("--conf-thres", type=float, default=0.25, help="confidence threshold")
    parser.add_argument("--iou-thres", type=float, default=0.45, help="NMS IoU threshold")
    parser.add_argument("--max-det", type=int, default=1000, help="maximum detections per image")
    parser.add_argument("--device", default="", help="cuda device, i.e. 0 or 0,1,2,3 or cpu")
    parser.add_argument("--view-img", action="store_true", help="show results")
    parser.add_argument("--save-txt", action="store_true", help="save results to *.txt")
    parser.add_argument("--save-csv", action="store_true", help="save results in CSV format")
    parser.add_argument("--save-conf", action="store_true", help="save confidences in --save-txt labels")
    parser.add_argument("--save-crop", action="store_true", help="save cropped prediction boxes")
    parser.add_argument("--nosave", action="store_true", help="do not save images/videos")
    parser.add_argument("--classes", nargs="+", type=int, help="filter by class: --classes 0, or --classes 0 2 3")
    parser.add_argument("--agnostic-nms", action="store_true", help="class-agnostic NMS")
    parser.add_argument("--augment", action="store_true", help="augmented inference")
    parser.add_argument("--visualize", action="store_true", help="visualize features")
    parser.add_argument("--update", action="store_true", help="update all models")
    parser.add_argument("--project", default=ROOT / "runs/detect", help="save results to project/name")
    parser.add_argument("--name", default="exp", help="save results to project/name")
    parser.add_argument("--exist-ok", action="store_true", help="existing project/name ok, do not increment")
    parser.add_argument("--line-thickness", default=3, type=int, help="bounding box thickness (pixels)")
    parser.add_argument("--hide-labels", default=False, action="store_true", help="hide labels")
    parser.add_argument("--hide-conf", default=False, action="store_true", help="hide confidences")
    parser.add_argument("--half", action="store_true", help="use FP16 half-precision inference")
    parser.add_argument("--dnn", action="store_true", help="use OpenCV DNN for ONNX inference")
    parser.add_argument("--vid-stride", type=int, default=1, help="video frame-rate stride")
    opt = parser.parse_args()
    opt.imgsz *= 2 if len(opt.imgsz) == 1 else 1  # expand
    print_args(vars(opt))
    return opt


def main(opt):
    check_requirements(ROOT / "requirements.txt", exclude=("tensorboard", "thop"))
    run(**vars(opt))


if __name__ == "__main__":
    opt = parse_opt()
    main(opt)
